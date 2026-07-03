"""
Supplier self-service portal views.
All views are gated by is_supplier() and scope every queryset to the
logged-in supplier via request.user.supplier_account.
"""
import re
import json as json_mod
from decimal import Decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncMonth
from django.utils import timezone
from django.utils.text import slugify
from django.core.paginator import Paginator

from apps.accounts.models import User, SiteSettings
from apps.vehicles.models import (
    Vehicle, VehicleCategory, VehicleFeature, VehicleZonePricing, VehicleImage, Supplier
)
from apps.locations.models import (
    Zone, ZoneDistanceRange, Route, VehicleRoutePricing,
    RoutePickupZone, RouteDropoffZone,
)
from apps.transfers.models import Transfer


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

SUPPLIER_LOGIN_URL = '/supplier/login/'


def is_supplier(user):
    return (
        user.is_authenticated
        and user.role == 'supplier'
        and getattr(user, 'supplier_account', None) is not None
    )


def supplier_required(view_func):
    """Decorator: require supplier login, redirect to supplier login page."""
    decorated = login_required(login_url=SUPPLIER_LOGIN_URL)(
        user_passes_test(is_supplier, login_url=SUPPLIER_LOGIN_URL)(view_func)
    )
    return decorated


def _get_supplier(request):
    """Return the Supplier linked to the logged-in user."""
    return request.user.supplier_account


# ---------------------------------------------------------------------------
# Login / Logout
# ---------------------------------------------------------------------------

def supplier_login(request):
    if request.user.is_authenticated and is_supplier(request.user):
        return redirect('supplier:home')

    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request, email=email, password=password)
        if user is not None and is_supplier(user):
            login(request, user)
            return redirect(request.GET.get('next', 'supplier:home'))
        messages.error(request, 'Invalid credentials or insufficient permissions.')

    return render(request, 'supplier/login.html')


def supplier_logout(request):
    logout(request)
    return redirect('supplier:login')


# ---------------------------------------------------------------------------
# Home / Dashboard
# ---------------------------------------------------------------------------

@supplier_required
def home(request):
    supplier = _get_supplier(request)
    today = timezone.now().date()
    month_start = today.replace(day=1)

    qs = Transfer.objects.filter(supplier=supplier)
    qs_month = qs.filter(pickup_datetime__date__gte=month_start)

    total_bookings = qs.count()
    month_bookings = qs_month.count()

    total_cost_mad = qs.aggregate(s=Sum('cost'))['s'] or Decimal('0')
    month_cost_mad = qs_month.aggregate(s=Sum('cost'))['s'] or Decimal('0')
    total_revenue_eur = qs.filter(currency='EUR').aggregate(s=Sum('total_price'))['s'] or Decimal('0')
    month_revenue_eur = qs_month.filter(currency='EUR').aggregate(s=Sum('total_price'))['s'] or Decimal('0')

    status_counts = {}
    for s_val, s_label in Transfer.Status.choices:
        status_counts[s_label] = qs.filter(status=s_val).count()

    recent_bookings = qs.select_related('vehicle_category').order_by('-pickup_datetime')[:10]

    vehicle_count = supplier.vehicles.filter(is_active=True).count()
    zone_count = Zone.objects.filter(owner_supplier=supplier, is_active=True).count()
    route_count = Route.objects.filter(owner_supplier=supplier, is_active=True).count()

    context = {
        'supplier': supplier,
        'total_bookings': total_bookings,
        'month_bookings': month_bookings,
        'total_cost_mad': total_cost_mad,
        'month_cost_mad': month_cost_mad,
        'total_revenue_eur': total_revenue_eur,
        'month_revenue_eur': month_revenue_eur,
        'status_counts': status_counts,
        'recent_bookings': recent_bookings,
        'vehicle_count': vehicle_count,
        'zone_count': zone_count,
        'route_count': route_count,
    }
    return render(request, 'supplier/home.html', context)


# ---------------------------------------------------------------------------
# Vehicles
# ---------------------------------------------------------------------------

@supplier_required
def vehicle_list(request):
    supplier = _get_supplier(request)
    vehicles = (
        supplier.vehicles
        .select_related('category')
        .prefetch_related('images')
        .filter(service_type='transfer')
        .order_by('name')
    )

    status_filter = request.GET.get('status')
    if status_filter:
        vehicles = vehicles.filter(status=status_filter)

    context = {
        'vehicles': vehicles,
        'statuses': Vehicle.Status.choices,
        'status_filter': status_filter or '',
    }
    return render(request, 'supplier/vehicles/list.html', context)


@supplier_required
def vehicle_create(request):
    supplier = _get_supplier(request)

    if request.method == 'POST':
        category_id = request.POST.get('category')
        name = request.POST.get('name', '').strip()
        passengers = request.POST.get('passengers')
        luggage = request.POST.get('luggage')

        if all([category_id, name, passengers, luggage]):
            try:
                custom_info = {}
                try:
                    custom_info = json_mod.loads(request.POST.get('custom_info', '{}') or '{}')
                except (json_mod.JSONDecodeError, ValueError):
                    pass

                key_features = []
                try:
                    key_features = json_mod.loads(request.POST.get('key_features', '[]') or '[]')
                except (json_mod.JSONDecodeError, ValueError):
                    pass

                vehicle = Vehicle.objects.create(
                    category_id=category_id,
                    name=name,
                    passengers=passengers,
                    luggage=luggage,
                    supplier=supplier,
                    supplier_name=supplier.name,
                    supplier_email=supplier.email,
                    service_type='transfer',
                    client_description=request.POST.get('client_description', ''),
                    key_features=key_features,
                    important_note=request.POST.get('important_note', ''),
                    important_note_type=request.POST.get('important_note_type', 'info'),
                    custom_info=custom_info,
                    status='available',
                    is_active=True,
                )

                if request.FILES.get('image'):
                    VehicleImage.objects.create(
                        vehicle=vehicle,
                        image=request.FILES['image'],
                        is_primary=True,
                    )

                # Zone pricing (from inline rows on create form)
                for key, value in request.POST.items():
                    if key.startswith('zone_price_') and value:
                        range_id = key.replace('zone_price_', '')
                        cost_val = request.POST.get(f'zone_cost_{range_id}') or None
                        min_hours = request.POST.get(f'zone_min_hours_{range_id}') or None
                        try:
                            VehicleZonePricing.objects.create(
                                vehicle=vehicle,
                                zone_distance_range_id=range_id,
                                price=value,
                                cost=cost_val,
                                min_booking_hours=min_hours,
                                is_active=True,
                            )
                        except Exception:
                            pass

                # Route pricing
                route_price_re = re.compile(r'route_price_(\d+)$')
                for key, value in request.POST.items():
                    m = route_price_re.match(key)
                    if m and value:
                        route_id = m.group(1)
                        cost_val = request.POST.get(f'route_cost_{route_id}') or None
                        min_hours = request.POST.get(f'route_min_hours_{route_id}') or None
                        pickup_adjs, dropoff_adjs = {}, {}
                        for adj_key, adj_val in request.POST.items():
                            if adj_key.startswith(f'route_{route_id}_pickup_adj_') and adj_val:
                                pickup_adjs[adj_key.replace(f'route_{route_id}_pickup_adj_', '')] = str(adj_val)
                            elif adj_key.startswith(f'route_{route_id}_dropoff_adj_') and adj_val:
                                dropoff_adjs[adj_key.replace(f'route_{route_id}_dropoff_adj_', '')] = str(adj_val)
                        try:
                            VehicleRoutePricing.objects.create(
                                vehicle=vehicle,
                                route_id=route_id,
                                price=value,
                                cost=cost_val,
                                min_booking_hours=min_hours,
                                pickup_zone_adjustments=pickup_adjs,
                                dropoff_zone_adjustments=dropoff_adjs,
                                is_active=True,
                            )
                        except Exception:
                            pass

                messages.success(request, f'Vehicle "{name}" created.')
                return redirect('supplier:vehicle_detail', pk=vehicle.pk)
            except Exception as e:
                messages.error(request, f'Error creating vehicle: {e}')
        else:
            messages.error(request, 'Please fill in all required fields (category, name, passengers, luggage).')

    categories = VehicleCategory.objects.filter(is_active=True)
    # Zones and routes this supplier owns (for inline pricing on create)
    zones = Zone.objects.prefetch_related('distance_ranges').filter(
        owner_supplier=supplier, is_active=True
    )
    routes = Route.objects.filter(
        owner_supplier=supplier, is_active=True
    ).prefetch_related('pickup_zones', 'dropoff_zones').order_by('name')

    context = {
        'categories': categories,
        'zones': zones,
        'routes': routes,
    }
    return render(request, 'supplier/vehicles/create.html', context)


@supplier_required
def vehicle_detail(request, pk):
    supplier = _get_supplier(request)
    vehicle = get_object_or_404(
        Vehicle.objects.select_related('category').prefetch_related(
            'images', 'zone_pricing__zone_distance_range__zone'
        ),
        pk=pk,
        supplier=supplier,
    )

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_vehicle':
            vehicle.category_id = request.POST.get('category')
            vehicle.name = request.POST.get('name', vehicle.name).strip()
            vehicle.passengers = request.POST.get('passengers', vehicle.passengers)
            vehicle.luggage = request.POST.get('luggage', vehicle.luggage)
            vehicle.status = request.POST.get('status', vehicle.status)
            vehicle.is_active = request.POST.get('is_active') == 'on'
            vehicle.client_description = request.POST.get('client_description', '')
            vehicle.notes = request.POST.get('notes', '')
            vehicle.important_note = request.POST.get('important_note', '')
            vehicle.important_note_type = request.POST.get('important_note_type', 'info')
            try:
                vehicle.key_features = json_mod.loads(request.POST.get('key_features', '[]') or '[]')
            except (json_mod.JSONDecodeError, ValueError):
                pass
            try:
                vehicle.custom_info = json_mod.loads(request.POST.get('custom_info', '{}') or '{}')
            except (json_mod.JSONDecodeError, ValueError):
                pass
            vehicle.save()
            messages.success(request, 'Vehicle updated.')

        elif action == 'add_zone_pricing':
            range_id = request.POST.get('zone_distance_range')
            price = request.POST.get('price')
            if range_id and price:
                try:
                    VehicleZonePricing.objects.create(
                        vehicle=vehicle,
                        zone_distance_range_id=range_id,
                        price=price,
                        cost=request.POST.get('cost') or None,
                        min_booking_hours=request.POST.get('min_booking_hours') or None,
                        is_active=True,
                    )
                    messages.success(request, 'Zone pricing added.')
                except Exception as e:
                    messages.error(request, f'Error: {e}')
            else:
                messages.error(request, 'Zone range and price are required.')

        elif action == 'update_zone_pricing':
            pricing = get_object_or_404(VehicleZonePricing, pk=request.POST.get('pricing_id'), vehicle=vehicle)
            pricing.price = request.POST.get('price', pricing.price)
            pricing.cost = request.POST.get('cost') or None
            pricing.min_booking_hours = request.POST.get('min_booking_hours') or None
            pricing.save()
            messages.success(request, 'Zone pricing updated.')

        elif action == 'delete_zone_pricing':
            pricing = get_object_or_404(VehicleZonePricing, pk=request.POST.get('pricing_id'), vehicle=vehicle)
            pricing.delete()
            messages.success(request, 'Zone pricing deleted.')

        elif action == 'add_route_pricing':
            route_id = request.POST.get('route_id')
            price = request.POST.get('price')
            if route_id and price:
                try:
                    VehicleRoutePricing.objects.create(
                        vehicle=vehicle,
                        route_id=route_id,
                        price=price,
                        cost=request.POST.get('cost') or None,
                        min_booking_hours=request.POST.get('min_booking_hours') or None,
                        is_active=True,
                    )
                    messages.success(request, 'Route pricing added.')
                except Exception as e:
                    messages.error(request, f'Error: {e}')
            else:
                messages.error(request, 'Route and price are required.')

        elif action == 'update_route_pricing':
            pricing = get_object_or_404(VehicleRoutePricing, pk=request.POST.get('pricing_id'), vehicle=vehicle)
            pricing.price = request.POST.get('price', pricing.price)
            pricing.cost = request.POST.get('cost') or None
            pricing.min_booking_hours = request.POST.get('min_booking_hours') or None
            pickup_adjs, dropoff_adjs = {}, {}
            for key, val in request.POST.items():
                if key.startswith('pickup_adj_') and val:
                    pickup_adjs[key.replace('pickup_adj_', '')] = str(val)
                elif key.startswith('dropoff_adj_') and val:
                    dropoff_adjs[key.replace('dropoff_adj_', '')] = str(val)
            if pickup_adjs:
                pricing.pickup_zone_adjustments = pickup_adjs
            if dropoff_adjs:
                pricing.dropoff_zone_adjustments = dropoff_adjs
            pricing.save()
            messages.success(request, 'Route pricing updated.')

        elif action == 'delete_route_pricing':
            pricing = get_object_or_404(VehicleRoutePricing, pk=request.POST.get('pricing_id'), vehicle=vehicle)
            pricing.delete()
            messages.success(request, 'Route pricing deleted.')

        elif action == 'upload_image':
            img = request.FILES.get('image')
            if img:
                VehicleImage.objects.create(
                    vehicle=vehicle,
                    image=img,
                    is_primary=request.POST.get('is_primary') == 'on',
                )
                messages.success(request, 'Image uploaded.')
            else:
                messages.error(request, 'Please select an image.')

        elif action == 'delete_image':
            img = get_object_or_404(VehicleImage, pk=request.POST.get('image_id'), vehicle=vehicle)
            img.delete()
            messages.success(request, 'Image deleted.')

        elif action == 'set_primary_image':
            img = get_object_or_404(VehicleImage, pk=request.POST.get('image_id'), vehicle=vehicle)
            img.is_primary = True
            img.save()
            messages.success(request, 'Primary image updated.')

        elif action == 'delete_vehicle':
            name = vehicle.name
            vehicle.delete()
            messages.success(request, f'Vehicle "{name}" deleted.')
            return redirect('supplier:vehicle_list')

        return redirect('supplier:vehicle_detail', pk=pk)

    categories = VehicleCategory.objects.filter(is_active=True)

    zone_pricing = vehicle.zone_pricing.select_related(
        'zone_distance_range__zone'
    ).filter(is_active=True).order_by('zone_distance_range__zone__name', 'zone_distance_range__min_km')

    existing_range_ids = vehicle.zone_pricing.values_list('zone_distance_range_id', flat=True)
    # Show ranges from this supplier's zones
    supplier_zone_ranges = ZoneDistanceRange.objects.select_related('zone').filter(
        is_active=True,
        zone__owner_supplier=supplier,
    ).order_by('zone__name', 'min_km')
    available_ranges = supplier_zone_ranges.exclude(id__in=existing_range_ids)

    route_pricing = vehicle.route_pricing.select_related('route').filter(
        is_active=True
    ).order_by('route__name')
    for rp in route_pricing:
        pickup_zones = rp.route.pickup_zones.filter(is_active=True).order_by('order', 'name')
        dropoff_zones = rp.route.dropoff_zones.filter(is_active=True).order_by('order', 'name')
        for z in pickup_zones:
            z.current_adjustment = rp.pickup_zone_adjustments.get(str(z.id), '0')
        for z in dropoff_zones:
            z.current_adjustment = rp.dropoff_zone_adjustments.get(str(z.id), '0')
        rp.route_pickup_zones = list(pickup_zones) if pickup_zones.exists() else None
        rp.route_dropoff_zones = list(dropoff_zones) if dropoff_zones.exists() else None

    existing_route_ids = set(vehicle.route_pricing.filter(is_active=True).values_list('route_id', flat=True))
    available_routes = Route.objects.filter(
        owner_supplier=supplier, is_active=True
    ).exclude(id__in=existing_route_ids).order_by('name')

    vehicle_images = vehicle.images.all().order_by('-is_primary', 'order', 'created_at')

    context = {
        'vehicle': vehicle,
        'categories': categories,
        'statuses': Vehicle.Status.choices,
        'zone_pricing': zone_pricing,
        'available_ranges': available_ranges,
        'route_pricing': route_pricing,
        'available_routes': available_routes,
        'vehicle_images': vehicle_images,
    }
    return render(request, 'supplier/vehicles/detail.html', context)


# ---------------------------------------------------------------------------
# Zones
# ---------------------------------------------------------------------------

@supplier_required
def zone_list(request):
    supplier = _get_supplier(request)
    zones = Zone.objects.prefetch_related('distance_ranges').filter(
        owner_supplier=supplier
    ).order_by('name')
    return render(request, 'supplier/zones/list.html', {'zones': zones})


@supplier_required
def zone_create(request):
    supplier = _get_supplier(request)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            slug = slugify(name)
            base_slug = slug
            counter = 1
            while Zone.objects.filter(slug=slug).exists():
                slug = f'{base_slug}-{counter}'
                counter += 1

            custom_info = {}
            try:
                custom_info = json_mod.loads(request.POST.get('custom_info', '{}') or '{}')
            except (json_mod.JSONDecodeError, ValueError):
                pass

            zone = Zone.objects.create(
                name=name,
                slug=slug,
                description=request.POST.get('description', ''),
                color=request.POST.get('color', '#3498db'),
                deposit_percentage=request.POST.get('deposit_percentage', 0) or 0,
                client_notice=request.POST.get('client_notice', ''),
                client_notice_type=request.POST.get('client_notice_type', 'info'),
                pickup_instructions=request.POST.get('pickup_instructions', ''),
                area_description=request.POST.get('area_description', ''),
                display_order=request.POST.get('display_order', 0) or 0,
                custom_info=custom_info,
                owner_supplier=supplier,
                is_active=True,
                extra_km_price=request.POST.get('extra_km_price', 0) or 0,
                max_extension_km=request.POST.get('max_extension_km', 0) or 0,
            )
            messages.success(request, f'Zone "{name}" created.')
            return redirect('supplier:zone_detail', pk=zone.pk)
        else:
            messages.error(request, 'Zone name is required.')

    return render(request, 'supplier/zones/create.html')


@supplier_required
def zone_detail(request, pk):
    supplier = _get_supplier(request)
    zone = get_object_or_404(
        Zone.objects.prefetch_related('distance_ranges'),
        pk=pk,
        owner_supplier=supplier,
    )

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_zone':
            new_max_extension_km = request.POST.get('max_extension_km', 0) or 0
            new_extra_km_price = request.POST.get('extra_km_price', 0) or 0
            # Hard block: reject save if the extension would conflict with a route endpoint
            if float(new_max_extension_km) > 0 and float(new_extra_km_price) > 0:
                from apps.locations.api.views import _check_zone_extension_conflicts
                conflicts = _check_zone_extension_conflicts(zone, new_max_extension_km)
                if conflicts:
                    names = ', '.join(f'"{c["route_name"]}" ({c["point"]}, {c["distance_km"]} km)' for c in conflicts)
                    messages.error(request, f'Cannot save: max extension km conflicts with route endpoint(s): {names}. Reduce the extension or contact support.')
                    return redirect('supplier:zone_detail', pk=pk)

            zone.name = request.POST.get('name', zone.name).strip() or zone.name
            zone.description = request.POST.get('description', '')
            zone.color = request.POST.get('color', zone.color)
            zone.deposit_percentage = request.POST.get('deposit_percentage', 0) or 0
            zone.client_notice = request.POST.get('client_notice', '')
            zone.client_notice_type = request.POST.get('client_notice_type', 'info')
            zone.pickup_instructions = request.POST.get('pickup_instructions', '')
            zone.area_description = request.POST.get('area_description', '')
            zone.is_active = request.POST.get('is_active') == 'on'
            zone.extra_km_price = new_extra_km_price
            zone.max_extension_km = new_max_extension_km
            try:
                zone.custom_info = json_mod.loads(request.POST.get('custom_info', '{}') or '{}')
            except (json_mod.JSONDecodeError, ValueError):
                pass
            zone.save()
            messages.success(request, 'Zone updated.')

        elif action == 'update_location':
            lat = request.POST.get('center_latitude')
            lng = request.POST.get('center_longitude')
            radius_km = request.POST.get('radius_km')
            if lat and lng:
                zone.center_latitude = lat
                zone.center_longitude = lng
            if radius_km:
                zone.radius_km = radius_km
            zone.save()
            messages.success(request, 'Zone location updated.')

        elif action == 'add_range':
            range_name = request.POST.get('range_name', '').strip()
            min_km = request.POST.get('min_km')
            max_km = request.POST.get('max_km')
            if all([range_name, min_km, max_km]):
                try:
                    dr = ZoneDistanceRange(zone=zone, name=range_name, min_km=min_km, max_km=max_km, is_active=True)
                    dr.full_clean()
                    dr.save()
                    messages.success(request, f'Distance range "{range_name}" added.')
                except Exception as e:
                    messages.error(request, str(e))
            else:
                messages.error(request, 'All fields are required for a distance range.')

        elif action == 'update_range':
            dr = get_object_or_404(ZoneDistanceRange, pk=request.POST.get('range_id'), zone=zone)
            dr.name = request.POST.get('range_name', dr.name)
            dr.min_km = request.POST.get('min_km', dr.min_km)
            dr.max_km = request.POST.get('max_km', dr.max_km)
            dr.is_active = request.POST.get('is_active') == 'on'
            try:
                dr.full_clean()
                dr.save()
                messages.success(request, 'Distance range updated.')
            except Exception as e:
                messages.error(request, str(e))

        elif action == 'delete_range':
            dr = get_object_or_404(ZoneDistanceRange, pk=request.POST.get('range_id'), zone=zone)
            dr.delete()
            messages.success(request, 'Distance range deleted.')

        elif action == 'delete_zone':
            name = zone.name
            zone.delete()
            messages.success(request, f'Zone "{name}" deleted.')
            return redirect('supplier:zone_list')

        return redirect('supplier:zone_detail', pk=pk)

    distance_ranges = zone.distance_ranges.all().order_by('min_km')
    site_settings = SiteSettings.get_settings()

    context = {
        'zone': zone,
        'distance_ranges': distance_ranges,
        'google_maps_api_key': site_settings.google_maps_api_key,
    }
    return render(request, 'supplier/zones/detail.html', context)


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@supplier_required
def route_list(request):
    supplier = _get_supplier(request)
    routes = Route.objects.filter(owner_supplier=supplier).annotate(
        vehicle_count=Count('vehicle_pricing')
    ).order_by('name')
    return render(request, 'supplier/routes/list.html', {'routes': routes})


@supplier_required
def route_create(request):
    supplier = _get_supplier(request)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        origin_name = request.POST.get('origin_name', '').strip()
        destination_name = request.POST.get('destination_name', '').strip()

        if all([name, origin_name, destination_name]):
            slug = slugify(name)
            base_slug = slug
            counter = 1
            while Route.objects.filter(slug=slug).exists():
                slug = f'{base_slug}-{counter}'
                counter += 1

            highlights = []
            try:
                highlights = json_mod.loads(request.POST.get('highlights', '[]') or '[]')
            except (json_mod.JSONDecodeError, ValueError):
                pass
            included_amenities = []
            try:
                included_amenities = json_mod.loads(request.POST.get('included_amenities', '[]') or '[]')
            except (json_mod.JSONDecodeError, ValueError):
                pass
            custom_info = {}
            try:
                custom_info = json_mod.loads(request.POST.get('custom_info', '{}') or '{}')
            except (json_mod.JSONDecodeError, ValueError):
                pass

            route = Route.objects.create(
                name=name,
                slug=slug,
                description=request.POST.get('description', ''),
                origin_name=origin_name,
                origin_latitude=request.POST.get('origin_latitude') or None,
                origin_longitude=request.POST.get('origin_longitude') or None,
                origin_radius_km=request.POST.get('origin_radius_km') or 10,
                destination_name=destination_name,
                destination_latitude=request.POST.get('destination_latitude') or None,
                destination_longitude=request.POST.get('destination_longitude') or None,
                destination_radius_km=request.POST.get('destination_radius_km') or 10,
                distance_km=request.POST.get('distance_km') or 0,
                estimated_duration_minutes=request.POST.get('estimated_duration_minutes') or None,
                deposit_percentage=request.POST.get('deposit_percentage', 0) or 0,
                client_notice=request.POST.get('client_notice', ''),
                client_notice_type=request.POST.get('client_notice_type', 'info'),
                route_description=request.POST.get('route_description', ''),
                travel_tips=request.POST.get('travel_tips', ''),
                estimated_traffic_info=request.POST.get('estimated_traffic_info', ''),
                highlights=highlights,
                included_amenities=included_amenities,
                custom_info=custom_info,
                is_bidirectional=True,
                is_active=request.POST.get('is_active') == 'on',
                owner_supplier=supplier,
            )
            messages.success(request, f'Route "{name}" created.')
            return redirect('supplier:route_detail', pk=route.pk)
        else:
            messages.error(request, 'Name, origin, and destination are required.')

    site_settings = SiteSettings.get_settings()
    return render(request, 'supplier/routes/create.html', {
        'GOOGLE_MAPS_API_KEY': site_settings.google_maps_api_key,
    })


@supplier_required
def route_detail(request, pk):
    supplier = _get_supplier(request)
    route = get_object_or_404(
        Route.objects.prefetch_related(
            'vehicle_pricing__vehicle__category',
            'vehicle_pricing__pickup_zone',
            'vehicle_pricing__dropoff_zone',
            'pickup_zones',
            'dropoff_zones',
        ),
        pk=pk,
        owner_supplier=supplier,
    )

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_route':
            route.name = request.POST.get('name', route.name).strip() or route.name
            route.description = request.POST.get('description', '')
            route.origin_name = request.POST.get('origin_name', route.origin_name)
            route.origin_latitude = request.POST.get('origin_latitude') or None
            route.origin_longitude = request.POST.get('origin_longitude') or None
            route.origin_radius_km = request.POST.get('origin_radius_km') or 10
            route.destination_name = request.POST.get('destination_name', route.destination_name)
            route.destination_latitude = request.POST.get('destination_latitude') or None
            route.destination_longitude = request.POST.get('destination_longitude') or None
            route.destination_radius_km = request.POST.get('destination_radius_km') or 10
            route.distance_km = request.POST.get('distance_km', route.distance_km)
            route.estimated_duration_minutes = request.POST.get('estimated_duration_minutes') or None
            route.deposit_percentage = request.POST.get('deposit_percentage', 0) or 0
            route.client_notice = request.POST.get('client_notice', '')
            route.client_notice_type = request.POST.get('client_notice_type', 'info')
            route.route_description = request.POST.get('route_description', '')
            route.travel_tips = request.POST.get('travel_tips', '')
            route.estimated_traffic_info = request.POST.get('estimated_traffic_info', '')
            try:
                route.highlights = json_mod.loads(request.POST.get('highlights', '[]') or '[]')
            except (json_mod.JSONDecodeError, ValueError):
                pass
            try:
                route.included_amenities = json_mod.loads(request.POST.get('included_amenities', '[]') or '[]')
            except (json_mod.JSONDecodeError, ValueError):
                pass
            try:
                route.custom_info = json_mod.loads(request.POST.get('custom_info', '{}') or '{}')
            except (json_mod.JSONDecodeError, ValueError):
                pass
            route.is_bidirectional = request.POST.get('is_bidirectional') == 'on'
            route.is_active = request.POST.get('is_active') == 'on'
            route.save()
            messages.success(request, 'Route updated.')

        elif action == 'add_pickup_zone':
            name = request.POST.get('zone_name', '').strip()
            lat = request.POST.get('latitude')
            lng = request.POST.get('longitude')
            if all([name, lat, lng]):
                if RoutePickupZone.objects.filter(route=route, name__iexact=name).exists():
                    messages.error(request, f'A pickup sub-zone named "{name}" already exists.')
                else:
                    RoutePickupZone.objects.create(
                        route=route, name=name,
                        center_latitude=lat, center_longitude=lng,
                        radius_km=request.POST.get('radius_km') or 5,
                        color=request.POST.get('color', '#28a745'),
                        price_adjustment=request.POST.get('price_adjustment') or 0,
                    )
                    messages.success(request, f'Pickup sub-zone "{name}" added.')
            else:
                messages.error(request, 'Name and coordinates are required.')

        elif action == 'update_pickup_zone':
            pz = get_object_or_404(RoutePickupZone, pk=request.POST.get('zone_id'), route=route)
            pz.name = request.POST.get('zone_name', pz.name)
            pz.center_latitude = request.POST.get('latitude') or pz.center_latitude
            pz.center_longitude = request.POST.get('longitude') or pz.center_longitude
            pz.radius_km = request.POST.get('radius_km') or pz.radius_km
            pz.color = request.POST.get('color') or pz.color
            if 'price_adjustment' in request.POST:
                pz.price_adjustment = request.POST.get('price_adjustment') or 0
            if 'is_active' in request.POST:
                pz.is_active = request.POST.get('is_active') == 'on'
            pz.save()
            messages.success(request, 'Pickup sub-zone updated.')

        elif action == 'delete_pickup_zone':
            pz = get_object_or_404(RoutePickupZone, pk=request.POST.get('zone_id'), route=route)
            pz.delete()
            messages.success(request, 'Pickup sub-zone deleted.')

        elif action == 'add_dropoff_zone':
            name = request.POST.get('zone_name', '').strip()
            lat = request.POST.get('latitude')
            lng = request.POST.get('longitude')
            if all([name, lat, lng]):
                if RouteDropoffZone.objects.filter(route=route, name__iexact=name).exists():
                    messages.error(request, f'A dropoff sub-zone named "{name}" already exists.')
                else:
                    RouteDropoffZone.objects.create(
                        route=route, name=name,
                        center_latitude=lat, center_longitude=lng,
                        radius_km=request.POST.get('radius_km') or 5,
                        color=request.POST.get('color', '#dc3545'),
                        price_adjustment=request.POST.get('price_adjustment') or 0,
                    )
                    messages.success(request, f'Dropoff sub-zone "{name}" added.')
            else:
                messages.error(request, 'Name and coordinates are required.')

        elif action == 'update_dropoff_zone':
            dz = get_object_or_404(RouteDropoffZone, pk=request.POST.get('zone_id'), route=route)
            dz.name = request.POST.get('zone_name', dz.name)
            dz.center_latitude = request.POST.get('latitude') or dz.center_latitude
            dz.center_longitude = request.POST.get('longitude') or dz.center_longitude
            dz.radius_km = request.POST.get('radius_km') or dz.radius_km
            dz.color = request.POST.get('color') or dz.color
            if 'price_adjustment' in request.POST:
                dz.price_adjustment = request.POST.get('price_adjustment') or 0
            if 'is_active' in request.POST:
                dz.is_active = request.POST.get('is_active') == 'on'
            dz.save()
            messages.success(request, 'Dropoff sub-zone updated.')

        elif action == 'delete_dropoff_zone':
            dz = get_object_or_404(RouteDropoffZone, pk=request.POST.get('zone_id'), route=route)
            dz.delete()
            messages.success(request, 'Dropoff sub-zone deleted.')

        elif action == 'add_vehicle_pricing':
            vehicle_id = request.POST.get('vehicle_id')
            price = request.POST.get('price')
            # Security: vehicle must belong to this supplier
            vehicle = get_object_or_404(Vehicle, pk=vehicle_id, supplier=supplier)
            if vehicle_id and price:
                try:
                    pickup_zone = None
                    dropoff_zone = None
                    pz_id = request.POST.get('pickup_zone_id')
                    dz_id = request.POST.get('dropoff_zone_id')
                    if pz_id:
                        pickup_zone = get_object_or_404(RoutePickupZone, pk=pz_id, route=route)
                    if dz_id:
                        dropoff_zone = get_object_or_404(RouteDropoffZone, pk=dz_id, route=route)
                    VehicleRoutePricing.objects.create(
                        vehicle=vehicle,
                        route=route,
                        pickup_zone=pickup_zone,
                        dropoff_zone=dropoff_zone,
                        price=price,
                        cost=request.POST.get('cost') or None,
                        min_booking_hours=request.POST.get('min_booking_hours') or None,
                        is_active=True,
                    )
                    messages.success(request, 'Vehicle pricing added.')
                except Exception as e:
                    messages.error(request, f'Error: {e}')
            else:
                messages.error(request, 'Vehicle and price are required.')

        elif action == 'update_vehicle_pricing':
            pricing = get_object_or_404(VehicleRoutePricing, pk=request.POST.get('pricing_id'), route=route)
            # Security: verify vehicle belongs to this supplier
            if pricing.vehicle.supplier_id != supplier.pk:
                messages.error(request, 'Not authorised.')
            else:
                pricing.price = request.POST.get('price', pricing.price)
                pricing.cost = request.POST.get('cost') or None
                pricing.min_booking_hours = request.POST.get('min_booking_hours') or None
                pricing.save()
                messages.success(request, 'Vehicle pricing updated.')

        elif action == 'delete_vehicle_pricing':
            pricing = get_object_or_404(VehicleRoutePricing, pk=request.POST.get('pricing_id'), route=route)
            if pricing.vehicle.supplier_id != supplier.pk:
                messages.error(request, 'Not authorised.')
            else:
                pricing.delete()
                messages.success(request, 'Vehicle pricing deleted.')

        elif action == 'delete_route':
            name = route.name
            route.delete()
            messages.success(request, f'Route "{name}" deleted.')
            return redirect('supplier:route_list')

        return redirect('supplier:route_detail', pk=pk)

    # Available vehicles for adding route pricing (this supplier's, not already priced on this route)
    priced_vehicle_ids = set(route.vehicle_pricing.values_list('vehicle_id', flat=True))
    available_vehicles = supplier.vehicles.filter(
        service_type='transfer', is_active=True
    ).exclude(id__in=priced_vehicle_ids).order_by('name')

    vehicle_pricing = route.vehicle_pricing.select_related('vehicle__category', 'pickup_zone', 'dropoff_zone').filter(
        vehicle__supplier=supplier,
    ).order_by('vehicle__name')

    site_settings = SiteSettings.get_settings()
    context = {
        'route': route,
        'pickup_zones': route.pickup_zones.filter(is_active=True).order_by('order', 'name'),
        'dropoff_zones': route.dropoff_zones.filter(is_active=True).order_by('order', 'name'),
        'vehicle_pricing': vehicle_pricing,
        'available_vehicles': available_vehicles,
        'GOOGLE_MAPS_API_KEY': site_settings.google_maps_api_key,
        'highlights_json': json_mod.dumps(route.highlights or []),
        'amenities_json': json_mod.dumps(route.included_amenities or []),
    }
    return render(request, 'supplier/routes/detail.html', context)


# ---------------------------------------------------------------------------
# Bookings
# ---------------------------------------------------------------------------

@supplier_required
def booking_list(request):
    supplier = _get_supplier(request)
    qs = Transfer.objects.filter(supplier=supplier).select_related(
        'vehicle_category', 'vehicle'
    ).order_by('-pickup_datetime')

    # Filters
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if status:
        qs = qs.filter(status=status)
    if date_from:
        qs = qs.filter(pickup_datetime__date__gte=date_from)
    if date_to:
        qs = qs.filter(pickup_datetime__date__lte=date_to)

    paginator = Paginator(qs, 25)
    page = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page,
        'transfers': page.object_list,
        'total_count': qs.count(),
        'statuses': Transfer.Status.choices,
        'status': status,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'supplier/bookings/list.html', context)


@supplier_required
def booking_detail(request, pk):
    supplier = _get_supplier(request)
    transfer = get_object_or_404(
        Transfer.objects.select_related('vehicle_category', 'vehicle', 'driver'),
        pk=pk,
        supplier=supplier,
    )

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'confirm' and transfer.status == Transfer.Status.PENDING:
            transfer.status = Transfer.Status.CONFIRMED
            transfer.save(update_fields=['status'])
            messages.success(request, 'Booking confirmed.')

        elif action == 'complete' and transfer.status in [
            Transfer.Status.CONFIRMED, Transfer.Status.PENDING
        ]:
            transfer.status = Transfer.Status.COMPLETED
            transfer.save(update_fields=['status'])
            messages.success(request, 'Booking marked as completed.')

        elif action == 'cancel' and transfer.status in [
            Transfer.Status.PENDING, Transfer.Status.CONFIRMED
        ]:
            transfer.status = Transfer.Status.CANCELLED
            transfer.save(update_fields=['status'])
            messages.success(request, 'Booking cancelled.')

        else:
            messages.error(request, 'Action not allowed for the current booking status.')

        return redirect('supplier:booking_detail', pk=pk)

    context = {
        'transfer': transfer,
    }
    return render(request, 'supplier/bookings/detail.html', context)


# ---------------------------------------------------------------------------
# Earnings
# ---------------------------------------------------------------------------

@supplier_required
def earnings(request):
    supplier = _get_supplier(request)
    today = timezone.now().date()
    month_start = today.replace(day=1)

    date_from_str = request.GET.get('date_from') or month_start.isoformat()
    date_to_str = request.GET.get('date_to') or today.isoformat()

    qs = Transfer.objects.filter(supplier=supplier).order_by('-pickup_datetime')
    if date_from_str:
        qs = qs.filter(pickup_datetime__date__gte=date_from_str)
    if date_to_str:
        qs = qs.filter(pickup_datetime__date__lte=date_to_str)

    if request.GET.get('format') == 'xlsx':
        return _earnings_xlsx(qs, supplier, date_from_str, date_to_str)

    total_cost_mad = qs.aggregate(s=Sum('cost'))['s'] or Decimal('0')
    total_revenue_eur = qs.filter(currency='EUR').aggregate(s=Sum('total_price'))['s'] or Decimal('0')
    total_bookings = qs.count()

    # Monthly breakdown
    monthly = (
        Transfer.objects.filter(supplier=supplier)
        .annotate(month=TruncMonth('pickup_datetime'))
        .values('month')
        .annotate(
            count=Count('id'),
            cost_mad=Sum('cost'),
            revenue_eur=Sum('total_price'),
        )
        .order_by('-month')[:12]
    )

    paginator = Paginator(qs, 25)
    page = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page,
        'transfers': page.object_list,
        'total_cost_mad': total_cost_mad,
        'total_revenue_eur': total_revenue_eur,
        'total_bookings': total_bookings,
        'monthly': monthly,
        'date_from': date_from_str,
        'date_to': date_to_str,
    }
    return render(request, 'supplier/earnings.html', context)


def _earnings_xlsx(qs, supplier, date_from_str, date_to_str):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = 'Earnings'

    headers = [
        'Booking Ref', 'Status', 'Pickup Date', 'Customer Name',
        'Pickup', 'Dropoff', 'Passengers', 'Vehicle Category',
        'Cost (MAD)', 'Price (EUR)', 'Currency',
    ]
    ws.append(headers)
    hfont = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='0d6efd')
    halign = Alignment(horizontal='center', vertical='center')
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=ci)
        c.font = hfont
        c.fill = hfill
        c.alignment = halign

    for t in qs:
        pd = timezone.localtime(t.pickup_datetime) if t.pickup_datetime else None
        ws.append([
            t.booking_ref,
            t.get_status_display(),
            pd.strftime('%Y-%m-%d %H:%M') if pd else '',
            t.customer_name or '',
            t.pickup_address or '',
            t.dropoff_address or '',
            t.passengers or '',
            t.vehicle_category.name if t.vehicle_category else '',
            float(t.cost) if t.cost else '',
            float(t.total_price) if t.total_price else '',
            t.currency or '',
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = (
        f'attachment; filename="earnings_{supplier.name.replace(" ", "_")}'
        f'_{date_from_str}_to_{date_to_str}.xlsx"'
    )
    return resp


# ---------------------------------------------------------------------------
# Account
# ---------------------------------------------------------------------------

@supplier_required
def account(request):
    supplier = _get_supplier(request)
    user = request.user

    if request.method == 'POST':
        action = request.POST.get('action', 'update_profile')

        if action == 'update_profile':
            # Update Supplier fields
            name = request.POST.get('name', '').strip()
            if name:
                supplier.name = name
            supplier.email = request.POST.get('email', supplier.email).strip()
            supplier.phone = request.POST.get('phone', supplier.phone).strip()
            supplier.notes = request.POST.get('notes', supplier.notes)
            supplier.save()
            # Keep User email in sync
            new_email = request.POST.get('email', '').strip()
            if new_email and new_email != user.email:
                if not User.objects.filter(email=new_email).exclude(pk=user.pk).exists():
                    user.email = new_email
                    user.save(update_fields=['email'])
                else:
                    messages.error(request, 'That email is already in use by another account.')
                    return redirect('supplier:account')
            messages.success(request, 'Profile updated.')

        elif action == 'change_password':
            old = request.POST.get('old_password', '')
            new1 = request.POST.get('new_password1', '')
            new2 = request.POST.get('new_password2', '')
            if not user.check_password(old):
                messages.error(request, 'Current password is incorrect.')
            elif new1 != new2:
                messages.error(request, 'New passwords do not match.')
            elif len(new1) < 8:
                messages.error(request, 'Password must be at least 8 characters.')
            else:
                user.set_password(new1)
                user.save()
                # Re-authenticate to avoid session invalidation
                login(request, user)
                messages.success(request, 'Password changed successfully.')

        return redirect('supplier:account')

    context = {
        'supplier': supplier,
        'user': user,
    }
    return render(request, 'supplier/account.html', context)
